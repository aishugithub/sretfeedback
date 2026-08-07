# ============================================================================
# verify_scoring.py  —  THE VERIFICATION GATE (spec Section 16, Phase 3)
# ============================================================================
# WHERE THIS FITS IN THE WHOLE APPLICATION
# ----------------------------------------------------------------------------
# The spec is emphatic (Sections 2 & 16): our scoring must reproduce the four
# approved "Feedback Report V2.0" workbooks EXACTLY. This script is the proof.
# It:
#   1. Reads the ORIGINAL responses straight out of each workbook's Sheet1
#      (the same raw answers the approved formulas were computed from),
#   2. Builds the question metadata from seed_data.py (single source of truth for
#      sections + scale weights), mapping workbook columns 1:1 to template
#      questions (the column order matches the template order by construction),
#   3. Feeds them through scoring.compute_scores() — the very function the live
#      report engine uses — and
#   4. Compares every section score AND the overall score to the value the
#      workbook itself computes (read with data_only=True), asserting they match
#      to a tight tolerance (1e-9).
#
# If this script prints ALL PASS, the engine is faithful to the approved
# formulas and the report exports built on top of it are trustworthy. Run it
# with:  python verify_scoring.py   (from the app/ folder).
# ----------------------------------------------------------------------------

import os
import openpyxl

import seed_data
import scoring


# ----------------------------------------------------------------------------
# Build a { scale_code -> options list } lookup from seed_data.SCALES so we can
# attach the correct options (labels + weights + fractions) to each question.
# ----------------------------------------------------------------------------
def _scale_options():
    out = {}
    for sc in seed_data.SCALES:
        out[sc["code"]] = [
            {"label": lbl, "weight": w, "fraction": fr, "display_order": order}
            for (lbl, w, fr, order) in sc["options"]
        ]
    return out


SCALE_OPTIONS = _scale_options()
FREE_TEXT_SCALES = {sc["code"] for sc in seed_data.SCALES if sc["is_free_text"]}


# ----------------------------------------------------------------------------
# Build the `questions` list (as scoring.compute_scores expects) for a category,
# straight from the seed template definition. Question IDs are just the 1-based
# position, which is also the display order and the workbook column offset.
# ----------------------------------------------------------------------------
def build_questions(category_code):
    tpl = next(t for t in seed_data.TEMPLATES if t["category_code"] == category_code)
    questions = []
    for i, (section, text, scale_code) in enumerate(tpl["questions"], start=1):
        questions.append({
            "id": i,
            "section": section,
            "text": text,
            "scale_code": scale_code,
            "is_free_text": scale_code in FREE_TEXT_SCALES,
            "display_order": i,
            "options": SCALE_OPTIONS.get(scale_code, []),
        })
    return questions


# ----------------------------------------------------------------------------
# Read the raw response rows out of a workbook's Sheet1. Data columns start at
# column B (index 2) and run 1:1 with the template questions. We stop at the
# first fully-blank row (the aggregation block begins after a blank separator),
# so we capture exactly the student rows the approved formula ranged over.
# The syllabus column is stored numerically (1 / 0.8 / 0.6) in the workbook; we
# pass values through untouched — scoring._coerce_fraction handles the number
# form, and the option-label form used by the live student app.
# ----------------------------------------------------------------------------
def read_responses(path, sheet_name, n_questions, n_data_rows):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    responses = []
    for row in range(2, 2 + n_data_rows):          # rows 2 .. (2+n-1)
        rec = {}
        any_value = False
        for qidx in range(1, n_questions + 1):     # question ids 1..n
            col = qidx + 1                          # +1 because col A is Timestamp
            val = ws.cell(row=row, column=col).value
            if val is not None and str(val).strip() != "":
                rec[qidx] = val
                any_value = True
        # Keep the row even if sparse, but skip an entirely empty trailing row.
        if any_value:
            responses.append(rec)
    return responses


# ----------------------------------------------------------------------------
# The four cases. For each: workbook path + sheet, the category, the number of
# scored+open questions, the data-row span the workbook formula used, and the
# EXPECTED section/overall values (read live from the workbook, below).
# ----------------------------------------------------------------------------
REPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                          "Report")

CASES = [
    # (label, filename, sheet, category, n_questions, n_data_rows,
    #    expected_cell_map {section_key/overall: (sheet, cell)})
    ("Theory", "Feedback Report V2.0 - T.xlsx", "Sheet1", "THEORY", 13, 105,
     {"syllabus": "C115", "faculty": "C116", "material": "C117",
      "exam": "C118", "overall": "C120"}),
    ("Lab", "Feedback Report V2.0 - L (1).xlsx", "Sheet 1", "LAB", 12, 68,
     {"syllabus": "C78", "faculty": "C79", "resources": "C80",
      "exam": "C81", "overall": "C83"}),
    ("Skill", "Feedback Report V2.0 - SL.xlsx", "Sheet1", "SKILL", 14, 68,
     {"syllabus": "C78", "faculty": "C79", "material": "C80",
      "assessment": "C81", "overall": "C83"}),
    ("AE", "Feedback Report V2.0 - AE.xlsx", "Sheet1", "AE", 9, 219,
     {"training": "C240", "material": "C241", "knowledge": "C242",
      "overall_sec": "C243", "overall": "C245"}),
]


def expected_values(path, cell_map):
    """Read the workbook's OWN computed section/overall values (data_only)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    # All target cells live on Sheet1 / 'Sheet 1' (the compute sheet).
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb["Sheet 1"]
    return {k: ws[c].value for k, c in cell_map.items()}


def run():
    print("=" * 74)
    print("SCORING VERIFICATION — computed engine vs approved workbook formulas")
    print("=" * 74)
    all_ok = True
    TOL = 1e-9  # match to (well within) the last displayed decimal

    for (label, fname, sheet, cat, nq, nrows, cell_map) in CASES:
        path = os.path.join(REPORT_DIR, fname)
        questions = build_questions(cat)
        responses = read_responses(path, sheet, nq, nrows)
        report_key = {"THEORY": "T", "LAB": "L", "SKILL": "SL", "AE": "AE"}[cat]

        # Run OUR engine with the approved default (Discussed Late -> 0).
        result = scoring.compute_scores(report_key, questions, responses,
                                        discussed_late_weight=None)
        # Keep SECTION scores and the GRAND TOTAL cleanly separated, because the
        # AE report has a *section* literally named "Overall" (key 'overall',
        # cell C243) AND a grand total (also called Overall, cell C245). Mixing
        # them in one dict would let one clobber the other.
        sec_by_key = {s["key"]: s["score"] for s in result["section_scores"]}
        grand_total = result["overall"]

        exp = expected_values(path, cell_map)

        print(f"\n{label}  (n_recorded computed = {result['n_recorded']}, "
              f"responses read = {len(responses)})")
        for key, exp_cell in cell_map.items():
            # Resolve which computed number this workbook cell should equal:
            #   'overall'      -> the grand total (mean of section scores)
            #   'overall_sec'  -> the AE section whose key is 'overall' (C243)
            #   anything else  -> the section with that same key
            if key == "overall":
                gval = grand_total
            elif key == "overall_sec":
                gval = sec_by_key.get("overall")
            else:
                gval = sec_by_key.get(key)
            eval_ = exp[key]
            if gval is None or eval_ is None:
                ok = (gval is None and eval_ is None)
                diff = "n/a"
            else:
                diff = abs(gval - eval_)
                ok = diff <= TOL
            all_ok = all_ok and ok
            status = "OK " if ok else "XX "
            gstr = "None" if gval is None else f"{gval:.10f}"
            estr = "None" if eval_ is None else f"{eval_:.10f}"
            print(f"   [{status}] {key:12s} computed={gstr}  workbook={estr}"
                  f"  Δ={diff}")

    print("\n" + "=" * 74)
    print("RESULT:", "ALL PASS ✔  — engine reproduces every approved formula"
          if all_ok else "FAILURES ✘ — see XX lines above")
    print("=" * 74)
    return all_ok


if __name__ == "__main__":
    ok = run()
    raise SystemExit(0 if ok else 1)
