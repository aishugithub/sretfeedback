# ============================================================================
# student_importer.py  —  Roster upload: ranges -> per-cycle students (spec v3 §7.1)
# ============================================================================
# WHERE THIS FITS IN THE WHOLE APPLICATION
# ----------------------------------------------------------------------------
# Under v3 the system "knows nothing it was not told" (§2). Every cycle begins
# with TWO uploads; this module handles the first — the STUDENT ROSTER ("for
# whom"). The admin uploads ~30-40 compact ROWS, each naming a contiguous
# register-number range for one (programme, year, section); this module:
#
#   1. parses the workbook (a "Roster" sheet of ranges + an optional
#      "Exclusions" sheet of individual register numbers already gone, §7.1),
#   2. validates everything BEFORE writing anything (a dry run), producing a
#      reconciliation summary the admin confirms ("11 programmes · 4 year-groups
#      · 1,847 students"), and
#   3. on confirmation, EXPANDS each range into individual `students` rows for
#      the cycle and records the source ranges in `roster_range`.
#
# The year of study is a STORED column supplied by the roster — never derived
# (v3 §2.1). The email is the only derivation: <reg_no>@sriher.edu.in (§4).
#
# REGISTER NUMBER FORMAT (spec §4):  E + PP + YY + RRR   e.g.  E0225014
#   E   constant prefix
#   PP  programme code digits (E02 -> programme 'E02')     — display/validation
#   YY  admission-year digits (25)                          — display only
#   RRR roll number (014)                                   — the part a range spans
# Expanding E0225001..E0225030 yields the 30 register numbers E0225001..E0225030.
# Out-of-range roll numbers (e.g. a lateral entrant given 201) are handled
# naturally: they are simply their own one-row range (§ "Roll number exceptions").
# ----------------------------------------------------------------------------

import re                            # register-number shape + prefix checks
from openpyxl import load_workbook, Workbook  # read .xlsx; write the template

from config import Config            # programme master + email domain
import services                      # email_for(), programme_meta()


# A register number is E + 2 programme digits + 2 year digits + 3+ roll digits.
# We allow 3+ roll digits so an out-of-range roll like 201 (3 digits) or a rare
# 4-digit roll still parses; the roll is "everything after the first 5 chars".
_REGNO_RE = re.compile(r"^E\d{2}\d{2}\d{2,}$")

# The six roster columns, in order (spec §7.1). Matched case-insensitively with
# loose spacing so a slightly-renamed header still loads.
ROSTER_HEADERS = ["programme_code", "year_of_study", "section",
                  "reg_no_start", "reg_no_end", "expected_count"]


def _norm(s):
    """Trim a cell to a clean string; None/blank -> ''. Used throughout."""
    return ("" if s is None else str(s)).strip()


def _split_regno(reg):
    """Split 'E0225014' into (prefix 'E0225', roll_int 14, roll_width 3).

    prefix = programme+year segment shared by a contiguous class range;
    roll    = the numeric tail a range spans. Returns (None, None, None) on a
    malformed register number so the caller can flag it.
    """
    if not _REGNO_RE.match(reg):
        return None, None, None
    prefix = reg[:5]          # 'E' + PP + YY  -> 'E0225'
    roll_str = reg[5:]        # '014' (or '201', or '0014')
    return prefix, int(roll_str), len(roll_str)


# ----------------------------------------------------------------------------
# parse_roster(path) — read the workbook into raw range dicts + an exclusions
# list. Sheet 1 (or a sheet whose name starts 'roster') holds the ranges; an
# optional sheet whose name starts 'exclusion' holds a single column of register
# numbers to omit up front (§7.1). Every row is tagged with its 1-based Excel row
# number so error messages point the admin at the exact line to fix.
# ----------------------------------------------------------------------------
def parse_roster(path):
    wb = load_workbook(path, read_only=True, data_only=True)

    # Locate the ranges sheet (named 'roster*' if present, else the first sheet)
    # and the optional exclusions sheet.
    ranges_ws = None
    excl_ws = None
    for ws in wb.worksheets:
        nm = (ws.title or "").strip().lower()
        if nm.startswith("exclusion"):
            excl_ws = ws
        elif ranges_ws is None and (nm.startswith("roster") or nm.startswith("sheet")):
            ranges_ws = ws
    if ranges_ws is None:
        ranges_ws = wb.worksheets[0]

    ranges = []
    header_seen = False
    for idx, row in enumerate(ranges_ws.iter_rows(values_only=True), start=1):
        if row is None or all(_norm(c) == "" for c in row):
            continue
        if not header_seen:      # first non-blank row is the header — skip it
            header_seen = True
            continue
        cells = list(row) + [None] * (6 - len(row))
        ranges.append({
            "excel_row": idx,
            "programme_code": _norm(cells[0]).upper(),
            "year_raw": _norm(cells[1]),
            "section": (_norm(cells[2]) or "NA").upper(),
            "reg_no_start": _norm(cells[3]).upper(),
            "reg_no_end": _norm(cells[4]).upper(),
            "expected_raw": _norm(cells[5]),
        })

    exclusions = []
    if excl_ws is not None:
        seen_header = False
        for row in excl_ws.iter_rows(values_only=True):
            if row is None or all(_norm(c) == "" for c in row):
                continue
            val = _norm(row[0]).upper()
            if not seen_header:  # skip the header cell
                seen_header = True
                # If the first cell already looks like a reg number, keep it.
                if _REGNO_RE.match(val):
                    exclusions.append(val)
                continue
            if val:
                exclusions.append(val)

    wb.close()
    return ranges, exclusions


# ----------------------------------------------------------------------------
# validate_and_expand(ranges, exclusions, master, cycle_code) — the dry run.
# Runs every §7.1 validation and, if no hard error, expands the ranges into the
# concrete student list. Returns (students, roster_rows, errors, warnings, recon)
# where `recon` is the reconciliation summary the admin confirms before commit.
#
# Validations (spec §7.1):
#   * programme_code exists in the programme master                 -> error
#   * the PP digits inside reg_no_start match programme_code        -> error
#   * start/end share the same programme+admission-year prefix      -> error
#   * reg_no_end >= reg_no_start                                     -> error
#   * year_of_study is 1..4                                         -> error
#   * expected_count == computed range size                        -> warning
#   * no overlapping ranges across rows                            -> error
#   * no duplicate register number after expansion                 -> error
# Nothing is written on any hard error (validate before writing).
# ----------------------------------------------------------------------------
def validate_and_expand(ranges, exclusions, master, cycle_code):
    errors, warnings = [], []
    students = {}        # reg_no -> student dict (dedupe + overlap detection)
    roster_rows = []     # the roster_range rows to persist
    reg_source = {}      # reg_no -> excel_row that first produced it (overlap msg)

    valid_progs = {r["code"] for r in master.execute("SELECT code FROM programme")}
    excl_set = {e for e in exclusions if e}

    for r in ranges:
        rn = r["excel_row"]
        prog = r["programme_code"]
        start, end = r["reg_no_start"], r["reg_no_end"]

        # programme_code must be real
        if prog not in valid_progs:
            errors.append(f"Row {rn}: programme_code '{prog}' is not in the "
                          f"programme master.")

        # year_of_study must be 1..4
        year = None
        if not r["year_raw"].isdigit() or not (1 <= int(r["year_raw"]) <= 4):
            errors.append(f"Row {rn}: year_of_study '{r['year_raw']}' must be 1-4.")
        else:
            year = int(r["year_raw"])

        # register numbers must be well-formed
        p_start, roll_start, w_start = _split_regno(start)
        p_end, roll_end, w_end = _split_regno(end)
        if p_start is None:
            errors.append(f"Row {rn}: reg_no_start '{start}' is malformed "
                          f"(expected like E0225001).")
        if p_end is None:
            errors.append(f"Row {rn}: reg_no_end '{end}' is malformed "
                          f"(expected like E0225030).")
        if p_start is None or p_end is None or year is None:
            continue  # cannot expand this row; other errors already recorded

        # start & end must share the same programme+year prefix
        if p_start != p_end:
            errors.append(f"Row {rn}: start '{start}' and end '{end}' have "
                          f"different programme/year prefixes.")
            continue

        # the PP digits inside the register number must match programme_code
        # (deliberate redundancy — a mismatch catches a real error, §7.1)
        pp_in_reg = "E" + p_start[1:3]
        if prog in valid_progs and pp_in_reg != prog:
            errors.append(f"Row {rn}: reg number programme '{pp_in_reg}' does not "
                          f"match programme_code '{prog}'.")

        # end >= start
        if roll_end < roll_start:
            errors.append(f"Row {rn}: reg_no_end < reg_no_start.")
            continue

        # expected_count cross-check (soft — warn only)
        size = roll_end - roll_start + 1
        if r["expected_raw"].isdigit() and int(r["expected_raw"]) != size:
            warnings.append(f"Row {rn}: expected_count {r['expected_raw']} != "
                            f"range size {size}.")

        # EXPAND the range into concrete students, honouring exclusions and
        # detecting overlaps with earlier rows.
        width = max(w_start, 3)
        prog_level = Config.PROGRAMMES.get(prog, (None, None))[1]
        for roll in range(roll_start, roll_end + 1):
            reg = f"{p_start}{roll:0{width}d}"
            if reg in students:
                errors.append(f"Row {rn}: register number '{reg}' overlaps the "
                              f"range on row {reg_source[reg]}.")
                continue
            reg_source[reg] = rn
            status = "excluded" if reg in excl_set else "active"
            students[reg] = {
                "reg_no": reg,
                "cycle_code": cycle_code,
                "name": None,                       # names not required (§7.1)
                "email": services.email_for(reg),   # the only derivation (§4)
                "dept_code": prog,
                "programme": prog_level,
                "year_of_study": year,
                "section": r["section"],
                "status": status,
            }

        roster_rows.append({
            "cycle_code": cycle_code,
            "programme_code": prog,
            "year_of_study": year,
            "section": r["section"],
            "reg_no_start": start,
            "reg_no_end": end,
            "expected_count": int(r["expected_raw"]) if r["expected_raw"].isdigit() else size,
        })

    if errors:
        students, roster_rows = {}, []   # write nothing on any hard error

    student_list = list(students.values())
    recon = _reconcile(student_list, excl_set)
    return student_list, roster_rows, errors, warnings, recon


# ----------------------------------------------------------------------------
# _reconcile(students, excl_set) — build the human confirmation summary the
# admin approves before commit (spec §7.1 "reconciliation summary"). Counts by
# programme and by year-group, plus the totals and how many are pre-excluded.
# ----------------------------------------------------------------------------
def _reconcile(students, excl_set):
    progs, years = {}, {}
    excluded = 0
    for s in students:
        progs[s["dept_code"]] = progs.get(s["dept_code"], 0) + 1
        years[s["year_of_study"]] = years.get(s["year_of_study"], 0) + 1
        if s["status"] == "excluded":
            excluded += 1
    return {
        "total": len(students),
        "programmes": len(progs),
        "year_groups": len(years),
        "excluded": excluded,
        "by_programme": dict(sorted(progs.items())),
        "by_year": dict(sorted(years.items())),
    }


# ----------------------------------------------------------------------------
# commit_roster(students, roster_rows, master, cycle_code) — write the expanded
# roster for the cycle. REPLACES any existing roster for this cycle first (a
# re-upload is authoritative), so the operation is idempotent and safe to redo.
# ----------------------------------------------------------------------------
def commit_roster(students, roster_rows, master, cycle_code):
    cur = master.cursor()
    # A fresh upload supersedes the previous roster for THIS cycle only.
    cur.execute("DELETE FROM students WHERE cycle_code = ?", (cycle_code,))
    cur.execute("DELETE FROM roster_range WHERE cycle_code = ?", (cycle_code,))
    for s in students:
        cur.execute(
            "INSERT INTO students (reg_no, cycle_code, name, email, dept_code, "
            "programme, year_of_study, section, status) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (s["reg_no"], s["cycle_code"], s["name"], s["email"], s["dept_code"],
             s["programme"], s["year_of_study"], s["section"], s["status"]),
        )
    for rr in roster_rows:
        cur.execute(
            "INSERT INTO roster_range (cycle_code, programme_code, year_of_study, "
            "section, reg_no_start, reg_no_end, expected_count) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (rr["cycle_code"], rr["programme_code"], rr["year_of_study"],
             rr["section"], rr["reg_no_start"], rr["reg_no_end"], rr["expected_count"]),
        )
    master.commit()
    return len(students)


# ----------------------------------------------------------------------------
# import_roster(path, master, cycle_code, commit=False) — the one call the admin
# route makes. commit=False is the DRY RUN (parse + validate + reconcile, write
# nothing); commit=True writes after the admin confirms the reconciliation.
# ----------------------------------------------------------------------------
def import_roster(path, master, cycle_code, commit=False):
    ranges, exclusions = parse_roster(path)
    students, roster_rows, errors, warnings, recon = validate_and_expand(
        ranges, exclusions, master, cycle_code)
    inserted = 0
    if commit and not errors:
        inserted = commit_roster(students, roster_rows, master, cycle_code)
    return {
        "range_rows": len(ranges),
        "exclusions": len(exclusions),
        "students": len(students),
        "inserted": inserted,
        "committed": bool(commit and not errors),
        "errors": errors,
        "warnings": warnings,
        "reconciliation": recon,
    }


# ----------------------------------------------------------------------------
# build_roster_template(path, cycle_label) — write the downloadable roster
# template (spec §7: templates are downloaded from within the app so there is
# one canonical version that always matches the parser). Sheet 1 = headers + one
# example row; Sheet 2 = instructions; Sheet 3 = the exclusions sheet stub.
# ----------------------------------------------------------------------------
def build_roster_template(path, cycle_label="AY 2026-27 CA1"):
    wb = Workbook()

    ws = wb.active
    ws.title = "Roster"
    ws.append(ROSTER_HEADERS)
    # One worked example row the admin overwrites.
    ws.append(["E02", 2, "A", "E0225001", "E0225030", 30])
    ws.append(["E02", 2, "B", "E0225031", "E0225060", 30])
    ws.append(["E71", 1, "NA", "E7126001", "E7126018", 18])

    ins = wb.create_sheet("Instructions")
    for line in [
        [f"Student roster template — {cycle_label}"],
        [""],
        ["One row per contiguous register-number range for a class."],
        ["programme_code : E01..E81 — must exist in the programme master."],
        ["year_of_study  : 1-4 — the class year. SUPPLIED here, never computed."],
        ["section        : A / B ... or NA if the class is not divided."],
        ["reg_no_start   : full first register number, e.g. E0225001."],
        ["reg_no_end     : full last register number, e.g. E0225030."],
        ["expected_count : the class size (cross-checked against the range)."],
        [""],
        ["Out-of-range roll numbers (e.g. a lateral entrant given 201) are just"],
        ["their own one-row range: E0225201 .. E0225201, expected_count 1."],
        [""],
        ["Exclusions sheet: list register numbers of students who have already"],
        ["left (TC/withdrawn) BEFORE the cycle. They are kept in the roster but"],
        ["marked excluded — no email, not counted in the expected totals."],
    ]:
        ins.append(line)

    excl = wb.create_sheet("Exclusions")
    excl.append(["register_no"])
    excl.append(["E0225099"])

    wb.save(path)
    return path


# ----------------------------------------------------------------------------
# generate_demo_roster(master, cycle_code, per_group=6) — TEST DATA ONLY.
# ----------------------------------------------------------------------------
# Synthesises a small roster for every (dept, year, section) group that appears
# in the cycle's offering allocation, so the student flow and readiness check can
# be demonstrated end-to-end without a real roster file. Deterministic register
# numbers keep it idempotent; the 'DEMO' prefix keeps demo data unmistakable.
# NEVER used in a production run.
# ----------------------------------------------------------------------------
def generate_demo_roster(master, cycle_code, per_group=6):
    groups = master.execute(
        "SELECT DISTINCT dept_code, year_of_study, section FROM offering "
        "WHERE cycle_code = ? ORDER BY dept_code, year_of_study, section",
        (cycle_code,),
    ).fetchall()
    made = 0
    cur = master.cursor()
    for g in groups:
        sect = g["section"] or "NA"
        for i in range(1, per_group + 1):
            reg = f"DEMO{g['dept_code']}{g['year_of_study']}{sect}{i:02d}"
            exists = cur.execute(
                "SELECT 1 FROM students WHERE reg_no = ? AND cycle_code = ?",
                (reg, cycle_code)).fetchone()
            if exists:
                continue
            cur.execute(
                "INSERT INTO students (reg_no, cycle_code, name, email, dept_code, "
                "programme, year_of_study, section, status) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'active')",
                (reg, cycle_code, f"Demo {reg}", f"{reg.lower()}@demo.sret.local",
                 g["dept_code"], Config.PROGRAMMES.get(g["dept_code"], (None, None))[1],
                 g["year_of_study"], sect),
            )
            made += 1
    master.commit()
    return made
