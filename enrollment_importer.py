# ============================================================================
# enrollment_importer.py  —  Elective enrollment upload (spec v3 §7.3)
# ============================================================================
# WHERE THIS FITS IN THE WHOLE APPLICATION
# ----------------------------------------------------------------------------
# CORE / LAB / PROJECT courses resolve their audience straight from the roster
# by (programme, year, section). An ELECTIVE cannot: one course code may be
# taught by two faculty to two different student sets, so the course code alone
# is not the key. The rigid allocation upload has already created a SEPARATE
# offering row per (elective, faculty), each with its own offering_id; this
# module attaches individual students to ONE of those rows.
#
# ADMIN FLOW (spec §7.3), enforced by the route that calls this module:
#   1. pick the elective course from a dropdown (populated from the allocation),
#   2. pick the faculty from a second dropdown showing ONLY faculty allocated to
#      that course — so which offering_id we attach to is CHOSEN, never typed,
#   3. upload a file that is a SINGLE COLUMN of register numbers.
#
# Validations (spec §7.3):
#   * every register number is present in THIS cycle's roster (the only
#     membership test — no year/programme/graduation logic, so M.Sc and
#     mixed-year electives work),
#   * a register number must not already be enrolled in the SAME course_code
#     under a different faculty (a student cannot be in both Dr A's and Dr B's
#     group), and
#   * duplicates within the file are collapsed.
# The reconciliation echoes "CS4001 — Dr A: 42, Dr B: 38, total 80".
# ----------------------------------------------------------------------------

from openpyxl import load_workbook, Workbook


def _norm(s):
    return ("" if s is None else str(s)).strip()


# ----------------------------------------------------------------------------
# parse_enrollment(path) — read the single-column register-number file. The
# first cell may be a header ('register_no'); anything that is not itself a
# register-number-looking token is treated as a header and skipped.
# ----------------------------------------------------------------------------
def parse_enrollment(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    regs = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row is None or all(_norm(c) == "" for c in row):
            continue
        val = _norm(row[0]).upper()
        if i == 0 and not val.startswith("E"):   # header cell like 'register_no'
            continue
        if val:
            regs.append(val)
    wb.close()
    return regs


# ----------------------------------------------------------------------------
# validate_enrollment(regs, master, cycle_code, offering_id) — the dry run.
# Returns (clean_regs, errors, warnings, recon). Writes nothing.
# ----------------------------------------------------------------------------
def validate_enrollment(regs, master, cycle_code, offering_id):
    errors, warnings = [], []

    off = master.execute(
        "SELECT * FROM offering WHERE id = ? AND cycle_code = ?",
        (offering_id, cycle_code)).fetchone()
    if off is None:
        return [], [f"Offering {offering_id} not found in cycle {cycle_code}."], [], {}
    if off["course_type"] != "ELECTIVE":
        return [], [f"Offering {offering_id} ({off['course_code']}) is not an "
                    f"ELECTIVE — enrollment applies to electives only."], [], {}

    # Roster membership set for the cycle.
    roster = {r["reg_no"] for r in master.execute(
        "SELECT reg_no FROM students WHERE cycle_code = ?", (cycle_code,))}

    # Students already enrolled in the SAME course_code under a DIFFERENT offering
    # (i.e. a different faculty's group) — a student can be in only one group.
    other_group = {}
    for r in master.execute(
            "SELECT e.reg_no, o.faculty FROM enrollment e "
            "JOIN offering o ON o.id = e.offering_id "
            "WHERE e.cycle_code = ? AND o.course_code = ? AND e.offering_id != ?",
            (cycle_code, off["course_code"], offering_id)):
        other_group[r["reg_no"]] = r["faculty"]

    seen, clean = set(), []
    for reg in regs:
        if reg in seen:
            warnings.append(f"'{reg}' listed more than once — counted once.")
            continue
        seen.add(reg)
        if reg not in roster:
            errors.append(f"'{reg}' is not in this cycle's roster.")
            continue
        if reg in other_group:
            errors.append(f"'{reg}' is already enrolled in {off['course_code']} "
                          f"under {other_group[reg]} — cannot be in two groups.")
            continue
        clean.append(reg)

    if errors:
        clean = []

    recon = {
        "course_code": off["course_code"],
        "faculty": off["faculty"],
        "offering_id": offering_id,
        "count": len(clean),
    }
    return clean, errors, warnings, recon


def commit_enrollment(clean, master, cycle_code, offering_id):
    """Replace this offering's enrollment with the validated set (idempotent)."""
    cur = master.cursor()
    cur.execute("DELETE FROM enrollment WHERE cycle_code = ? AND offering_id = ?",
                (cycle_code, offering_id))
    for reg in clean:
        cur.execute(
            "INSERT OR IGNORE INTO enrollment (cycle_code, offering_id, reg_no) "
            "VALUES (?, ?, ?)", (cycle_code, offering_id, reg))
    master.commit()
    return len(clean)


def import_enrollment(path, master, cycle_code, offering_id, commit=False):
    """The one call the admin route makes. commit=False is the dry run."""
    regs = parse_enrollment(path)
    clean, errors, warnings, recon = validate_enrollment(
        regs, master, cycle_code, offering_id)
    inserted = 0
    if commit and not errors:
        inserted = commit_enrollment(clean, master, cycle_code, offering_id)
    return {
        "rows": len(regs),
        "enrolled": len(clean),
        "inserted": inserted,
        "committed": bool(commit and not errors),
        "errors": errors,
        "warnings": warnings,
        "reconciliation": recon,
    }


def build_enrollment_template(path, course_code="CS4001", faculty="Dr. Anand"):
    """A trivial single-column template (spec §7.3)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Enrollment"
    ws.append(["register_no"])
    ws.append(["E0225001"])
    ws.append(["E0225002"])
    ins = wb.create_sheet("Instructions")
    ins.append([f"Enrollment for {course_code} — {faculty}"])
    ins.append(["One register number per row. Header 'register_no' optional."])
    ins.append(["Every student must be in this cycle's roster, and cannot appear"])
    ins.append(["in another faculty's group for the same course."])
    wb.save(path)
    return path
