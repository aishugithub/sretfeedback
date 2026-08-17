# ============================================================================
# report_export.py  —  Render a scored offering as Excel + print-ready PDF
#                      (spec Section 11 — the "Feedback Report V2.0" layout)
# ============================================================================
# WHERE THIS FITS IN THE WHOLE APPLICATION
# ----------------------------------------------------------------------------
# scoring.py turns anonymous answers into numbers. THIS module turns those
# numbers into the two documents the college actually hands out per course /
# faculty (spec Section 11):
#
#     build_excel_report(result, path)  -> an .xlsx matching Feedback Report V2.0
#     build_pdf_report(result, path)    -> a print-ready .pdf of the same
#
# Both consume the SAME `result` dict produced by scoring.score_offering(), so
# the Excel and the PDF can never disagree — they are two renderings of one
# computation. A small shared helper, `build_view_model(result)`, arranges the
# raw scores into the exact blocks the report shows, in order:
#     1. Header/identity block   (AY · Year/Term · Course · Programme · Faculty)
#     2. No. of Students Feedback Recorded
#     3. Overall Score            (the grand overall from scoring.py)
#     4. Section scores            (one row per report section)
#     5. Section tables            (per-question option counts + average /10)
#
# (v3.2) At the professor's request the report was SLIMMED DOWN — the per-section
# BAR CHARTS, the OPEN-ENDED COMMENTS block, and the DEPARTMENT-CODE LEGEND were
# all removed, leaving a compact, numbers-only document. build_view_model still
# computes the legend/comments fields (scoring is untouched); the renderers below
# simply no longer draw them.
#
# Bulk generation (spec Section 11 "bulk for a whole batch") is provided by
# build_batch_* helpers that loop these per-offering builders and package the
# output (a .zip of Excel files, or one multi-page combined PDF).
#
# Dependencies: openpyxl (already used for the importer) for Excel; reportlab
# (pure Python, pip-installable, no system libraries) for PDF — both run happily
# on the professor's laptop with no extra OS packages (spec Section 8/13).
# ----------------------------------------------------------------------------

import io
import os
import zipfile

# openpyxl — Excel workbook building. (v3.2: native bar charts removed with the
# report slim-down, so openpyxl.chart is no longer imported.)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# reportlab — PDF document building. (v3.2: the per-section bar chart was removed,
# so the reportlab.graphics chart/shape imports are no longer needed.)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, KeepTogether)

from config import Config


# ============================================================================
# SECTION 1 — THE DEPARTMENT-CODE LEGEND (spec Section 11)
# ============================================================================
# The approved report prints this legend verbatim under the header. We keep the
# full descriptive names here (the config map holds short names for validation);
# these long forms match the text block in the V2.0 workbooks (Sheet2!A18).
# ----------------------------------------------------------------------------
DEPT_LEGEND = [
    ("E01", "B. Tech Computer Science and Engineering (Artificial Intelligence and Machine Learning)"),
    ("E02", "B. Tech Computer Science and Engineering (Cyber Security and Internet of Things)"),
    ("E03", "B. Tech Computer Science and Engineering (Artificial Intelligence and Data Analytics)"),
    ("E05", "B. Tech Computer Science and Medical Engineering (Artificial Intelligence and Data Analytics)"),
    ("E06", "B. Tech Electronics and Communication Engineering"),
    ("E52", "B. Sc Computer Science (Artificial Intelligence and Data Analytics)"),
    ("E61", "B. Sc Bio Informatics"),
    ("E62", "B. Sc Data Science"),
    ("E71", "M. Sc Artificial Intelligence"),
    ("E73", "M. Sc Data Analytics"),
    ("E81", "M. Sc Medical Bioinformatics"),
]


# ============================================================================
# SECTION 2 — THE SHARED VIEW MODEL
# ============================================================================
# Arrange one scored offering's `result` into the ordered blocks both renderers
# draw. Keeping this in one place means the Excel and PDF are guaranteed to show
# the same identity line, the same section order, and the same tables.
# ----------------------------------------------------------------------------

# ----------------------------------------------------------------------------
# _name_titled(name) — move a trailing honorific to the FRONT (Aug 2026)
# ----------------------------------------------------------------------------
# The allocation stores teacher names with the title LAST, e.g. "Manoj Kumaran S
# Mr." or "Abhinand P A Dr.". Reports read better with the honorific in front —
# "Mr. Manoj Kumaran S", "Dr. Abhinand P A". We detect a final token that is a
# known honorific (with or without a dot, any case), normalise it to "Xx.", and
# move it to the front. Names without a trailing honorific are returned unchanged.
# ----------------------------------------------------------------------------
_HONORIFICS = {"dr": "Dr.", "mr": "Mr.", "ms": "Ms.", "mrs": "Mrs.",
               "prof": "Prof.", "miss": "Miss", "shri": "Shri", "smt": "Smt."}


def _name_titled(name):
    parts = (name or "").strip().split()
    if len(parts) < 2:
        return (name or "").strip()
    key = parts[-1].rstrip(".").lower()
    if key in _HONORIFICS:
        return _HONORIFICS[key] + " " + " ".join(parts[:-1])
    return " ".join(parts)


def _section_batch(section):
    """Section A -> 'Sec A / Batch 1', Section B -> 'Sec B / Batch 2', etc., so a
    report for one section of a split course clearly states which batch it covers.
    A course with no real section ('NA'/blank) returns '' (nothing to show)."""
    s = (section or "").strip().upper()
    if not s or s in ("NA", "N/A", "-", "NONE"):
        return ""
    idx = {"A": "1", "B": "2", "C": "3", "D": "4", "E": "5", "F": "6"}.get(s)
    return "Sec %s / Batch %s" % (s, idx) if idx else "Sec %s" % s


def build_view_model(result):
    o = result["offering"]

    # --- 2a. Header identity (mirrors workbook Sheet2!A4). -------------------
    # Year/Term is shown as "Y<n> - Sem <m>" using the offering's stored year and
    # semester. Missing pieces degrade gracefully to a blank rather than crash.
    year_term = f"Y{o['year_of_study']}"
    if o["semester"]:
        year_term += f" - Sem {o['semester']}"
    section = _section_batch(o["section"] if "section" in o.keys() else "")

    # ---- CONSOLIDATED-ELECTIVE header (Aug 2026 change) ----------------------
    # When this report was produced by scoring.score_offering_group over a pooled
    # elective delivery, `result["is_consolidated"]` is True and
    # `result["group_dept_codes"]` lists every programme code that sat in the one
    # class (e.g. ['E01','E02','E03']). In that case the PROGRAM CODE line shows
    # ALL of them and the PROGRAMME line names the elective basket it was pooled
    # across — instead of the anchor's single programme, which would misleadingly
    # read as if only that programme were surveyed. A normal (non-elective) report
    # keeps exactly its previous single-programme header.
    is_consolidated = bool(result.get("is_consolidated"))
    group_codes = result.get("group_dept_codes") or (
        [o["dept_code"]] if (o["dept_code"] or "") else [])
    basket = o["elective_basket"] if "elective_basket" in o.keys() else None
    if is_consolidated:
        dept_code_display = ", ".join(group_codes)
        programme_display = (
            "Elective basket %s — pooled across %s"
            % (("'%s'" % basket) if basket else "(elective)", ", ".join(group_codes)))
    else:
        # Show the OFFICIAL expanded programme name looked up by programme code
        # (Config.PROGRAMMES, kept current Aug 2026), falling back to the stored
        # value if the code is unknown.
        dept_code_display = o["dept_code"] or ""
        programme_display = (Config.PROGRAMMES.get(o["dept_code"] or "", (None,))[0]
                             or o["programme"] or "")

    header = {
        "section": section,
        "academic_year": o["academic_year"],
        "year_term": year_term,
        "course_code": o["course_code"] or "",
        "course_name": (o["course_name"] or "").upper(),
        "programme": programme_display,
        "dept_code": dept_code_display,
        "faculty": _name_titled(o["faculty"] or ""),
        "category": result.get("report_key") or "",
        "category_name": o["category_name"] if "category_name" in o.keys() else "",
        # Carried so a renderer could badge the report as a pooled elective; the
        # current layouts read it implicitly through the fields above.
        "is_consolidated": is_consolidated,
        "elective_basket": basket if is_consolidated else None,
    }

    # --- 2b. Section score rows (report section title + /10 score). ----------
    section_rows = [
        {"title": s["title"], "score": s["score"], "score_2dp": s["score_2dp"]}
        for s in result["section_scores"]
    ]

    # --- 2c. Per-section detail tables: for each report section, list its
    #         questions with the per-option counts, so we can draw the count
    #         tables + charts exactly as the workbook arranges them. Syllabus is
    #         a special case (a single fraction question) — we still show its
    #         option counts (100%/80%/60%).
    detail_sections = []
    pq = result["per_question"]
    # Index questions by id for quick lookup and to preserve display order.
    for s in result["section_scores"]:
        qblocks = []
        for qid in s["question_ids"]:
            info = pq.get(qid)
            if info is None or info["is_free_text"]:
                continue
            q = info["question"]
            # Ordered (label, count) pairs following the option display order.
            opts = sorted(q["options"], key=lambda op: op["display_order"])
            counts = [(op["label"], info["option_counts"].get(op["label"], 0))
                      for op in opts]
            qblocks.append({
                "text": q["text"],
                "counts": counts,
                "average": info["average"],
            })
        # Carry the SECTION score too. Some sections (notably Syllabus) score at
        # the section level and leave every per-question average None; the chart
        # falls back to this so such a section still shows a real bar.
        detail_sections.append({"title": s["title"], "questions": qblocks,
                                "score": s["score"]})

    return {
        "header": header,
        "legend": DEPT_LEGEND,
        "n_recorded": result["n_recorded"],
        "overall": result["overall"],
        "overall_2dp": result["overall_2dp"],
        "section_rows": section_rows,
        "detail_sections": detail_sections,
        "open_comments": result["open_comments"],
    }


def _fmt(v):
    """Format a /10 score for display: 2 decimals, or '—' when unanswered."""
    return "—" if v is None else f"{v:.2f}"


# ============================================================================
# SECTION 3 — EXCEL RENDERER (matches Feedback Report V2.0 layout)
# ============================================================================

# Reusable styling constants (kept here so a restyle is one-line-per-look).
_HDR_FILL = PatternFill("solid", fgColor="0B3D68")   # deep blue like the app
_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
_SUB_FILL = PatternFill("solid", fgColor="EEF1F4")
_BOLD = Font(bold=True)
_BIG = Font(bold=True, size=20, color="0B3D68")
_THIN = Side(style="thin", color="C9CED3")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def build_excel_report(result, path):
    """Write one offering's report to an .xlsx file at `path`.

    Layout order matches Section 11: header block, dept legend, recorded count,
    overall score, section-score table, then per-section count tables each with
    a bar chart, and finally the verbatim open comments.
    """
    vm = build_view_model(result)
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.sheet_view.showGridLines = False
    # Sensible column widths for a printable A4-ish sheet.
    widths = {"A": 46, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    r = 1  # running row cursor

    # ---- Banner image + title ----------------------------------------------
    # The college banner (app/static/banner.png) at the very top, then just
    # "Course Feedback Report" — no version, no "SRET" text. The image floats
    # over rows 1-7 (whose height we reserve); if it can't be added we simply
    # start the title at row 1 so the sheet is never broken.
    banner_path = os.path.join(Config.BASE_DIR, "static", "banner.png")
    if os.path.exists(banner_path):
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(banner_path)
            img.width, img.height = 720, 144          # 5:1 aspect, ~content width
            ws.add_image(img, "A1")
            for rr in range(1, 8):                    # reserve space under the image
                ws.row_dimensions[rr].height = 21
            r = 9
        except Exception:
            r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    cell = ws.cell(r, 1, "Course Feedback Report")
    cell.fill = _HDR_FILL; cell.font = _HDR_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 24
    r += 2

    # ---- Header identity block (label : value pairs) ------------------------
    h = vm["header"]
    ident = [
        ("ACADEMIC YEAR", h["academic_year"]),
        ("YEAR / TERM", h["year_term"]),
    ]
    if h["section"]:                       # only for split (multi-section) courses
        ident.append(("SECTION / BATCH", h["section"]))
    ident += [
        ("COURSE CODE", h["course_code"]),
        ("COURSE NAME", h["course_name"]),
        ("PROGRAM CODE", h["dept_code"]),
        ("PROGRAMME", h["programme"]),
        ("FACULTY NAME", h["faculty"]),
        ("CATEGORY", f"{h['category']} — {h['category_name']}"),
    ]
    for label, value in ident:
        ws.cell(r, 1, label).font = _BOLD
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        ws.cell(r, 2, value)
        r += 1
    r += 1

    # ---- No. of students + Overall score box --------------------------------
    ws.cell(r, 1, "No. of Students Feedback Recorded:").font = _BOLD
    ws.cell(r, 2, vm["n_recorded"])
    r += 1
    ws.cell(r, 1, "OVERALL SCORE (out of 10):").font = _BOLD
    oc = ws.cell(r, 2, vm["overall_2dp"] if vm["overall_2dp"] is not None else "—")
    oc.font = _BIG
    ws.row_dimensions[r].height = 28
    r += 2

    # ---- Section-score summary table ----------------------------------------
    ws.cell(r, 1, "Section").font = _HDR_FONT
    ws.cell(r, 1).fill = _HDR_FILL
    ws.cell(r, 2, "Score /10").font = _HDR_FONT
    ws.cell(r, 2).fill = _HDR_FILL
    r += 1
    for srow in vm["section_rows"]:
        ws.cell(r, 1, srow["title"]).border = _BORDER
        c = ws.cell(r, 2, srow["score_2dp"] if srow["score_2dp"] is not None else "—")
        c.border = _BORDER
        r += 1
    r += 1

    # ---- Per-section count tables -------------------------------------------
    # For each section we print a small table: rows = questions, columns = the
    # option labels, cells = counts, plus each question's average /10. (v3.2: the
    # clustered bar chart that used to follow each table was removed.)
    for sec in vm["detail_sections"]:
        if not sec["questions"]:
            continue
        ws.cell(r, 1, sec["title"]).font = _BIG
        ws.cell(r, 1).font = Font(bold=True, size=13, color="0B3D68")
        r += 1

        # Column headers = the union of option labels for this section's first
        # question (all questions in a section share one scale, so labels match).
        labels = [lbl for (lbl, _cnt) in sec["questions"][0]["counts"]]
        header_row = r
        ws.cell(r, 1, "Question").font = _BOLD
        ws.cell(r, 1).fill = _SUB_FILL
        for j, lbl in enumerate(labels, start=2):
            c = ws.cell(r, j, lbl)
            c.font = _BOLD; c.fill = _SUB_FILL
            c.alignment = Alignment(wrap_text=True, vertical="center")
        r += 1
        data_start = r
        for qb in sec["questions"]:
            ws.cell(r, 1, qb["text"]).alignment = Alignment(wrap_text=True)
            ws.cell(r, 1).border = _BORDER
            count_map = dict(qb["counts"])
            for j, lbl in enumerate(labels, start=2):
                c = ws.cell(r, j, count_map.get(lbl, 0))
                c.border = _BORDER
                c.alignment = Alignment(horizontal="center")
            r += 1
        data_end = r - 1
        r += 1   # a blank row before the next section's table

    # (v3.2) The department-code legend and the open-ended comments block that
    # used to close the sheet were removed here as part of the report slim-down.

    wb.save(path)
    return path


# ============================================================================
# SECTION 4 — PDF RENDERER (print-ready, matches the same layout)
# ============================================================================

def _pdf_styles():
    ss = getSampleStyleSheet()
    ss.add(ParagraphStyle("Ident", parent=ss["Normal"], fontSize=9, leading=12))
    ss.add(ParagraphStyle("SecTitle", parent=ss["Heading2"], textColor=colors.HexColor("#0B3D68")))
    ss.add(ParagraphStyle("Cmt", parent=ss["Normal"], fontSize=9, leading=12, leftIndent=8))
    ss.add(ParagraphStyle("Legend", parent=ss["Normal"], fontSize=7.5, leading=10))
    # Centred, small header style for the option-count table columns, so long
    # labels ("Strongly Disagree") wrap tidily inside their column instead of
    # overlapping the next one.
    ss.add(ParagraphStyle("Hdr", parent=ss["Normal"], fontSize=7, leading=8,
                          alignment=1, fontName="Helvetica-Bold"))
    return ss


# (v3.2) The per-question average bar chart (_section_avg_chart) and its bar
# colour constant were REMOVED here as part of the report slim-down. The exact
# per-question averages are still shown — as the numeric "Avg" column in each
# section table below — so no information is lost, only the graphic.


# ----------------------------------------------------------------------------
# NumberedCanvas — a reportlab canvas that stamps a proper footer with
# "Page X of Y" (the TOTAL is only known once every page is laid out).
#
# HOW IT WORKS: reportlab draws pages one at a time and cannot know the total up
# front. We override showPage()/save() to BUFFER each page's state; at save() the
# total is known, so we replay every page drawing the footer with the real X of Y.
#
# (v3.2) The report used to be PADDED to an even page count with a captioned
# duplex-pad page and a trailing blank "remarks" page (old spec §13.2). At the
# professor's request that padding was REMOVED — reports now print at their true
# length (no blank pages), so "Page X of Y" reflects real content pages only.
# ----------------------------------------------------------------------------
from reportlab.pdfgen import canvas as _rl_canvas  # noqa: E402


class NumberedCanvas(_rl_canvas.Canvas):
    footer_left = ""          # set per-build: "faculty · course_code"
    _watermark = None         # set per-build: diagonal text, e.g. "TEST DATA"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        # Buffer this page instead of emitting it, so we can revisit at save().
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        # Every buffered page is real content (v3.2: no blank padding is added).
        # The total is now known, so replay each page stamping the real "X of Y".
        total = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            if self._watermark:
                self._draw_watermark()
            self._draw_footer(total)
            super().showPage()
        super().save()

    def _draw_watermark(self):
        """Large, faint, diagonal watermark across the page (spec §9.1 test mode)."""
        self.saveState()
        self.setFont("Helvetica-Bold", 60)
        self.setFillColor(colors.Color(0.85, 0.2, 0.2, alpha=0.16))
        self.translate(A4[0] / 2.0, A4[1] / 2.0)
        self.rotate(45)
        self.drawCentredString(0, 0, self._watermark)
        self.restoreState()

    def _draw_footer(self, total):
        """Thin rule + three-zone footer on a numbered content page (spec §13.1)."""
        page_num = self._pageNumber
        self.saveState()
        y = 10 * mm
        self.setStrokeColor(colors.HexColor("#b8c2cc"))
        self.setLineWidth(0.5)
        self.line(15 * mm, y + 4 * mm, A4[0] - 15 * mm, y + 4 * mm)  # thin rule above
        self.setFont("Helvetica", 7)
        self.setFillColor(colors.HexColor("#555555"))
        self.drawString(15 * mm, y, self.footer_left[:80])
        self.drawCentredString(A4[0] / 2.0, y,
                               "Confidential — For faculty development purposes only")
        self.drawRightString(A4[0] - 15 * mm, y, f"Page {page_num} of {total}")
        self.restoreState()


def _footer_left_for(result):
    o = result["offering"]
    return f"{_name_titled(o['faculty']) or 'Faculty'} · {o['course_code'] or ''}"


def build_pdf_report(result, path_or_buffer):
    """Render one offering's report to a PDF (a filesystem path or a BytesIO,
    the latter used by the combined-batch PDF so pages can be concatenated).

    Uses NumberedCanvas so every page carries the confidential footer with
    "Page X of Y". (v3.2: the report is no longer padded — it prints at its true
    length with no trailing blank pages.)
    """
    ss = _pdf_styles()
    doc = SimpleDocTemplate(path_or_buffer, pagesize=A4,
                            topMargin=14 * mm, bottomMargin=18 * mm,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            title="SRET Feedback Report")
    story = []
    _append_pdf_story(result, story, ss)

    # Bind the per-report footer-left text onto the canvas class for this build.
    footer_left = _footer_left_for(result)
    watermark = result.get("watermark")   # e.g. "TEST DATA" for test cycles (§9.1)

    class _C(NumberedCanvas):
        pass
    _C.footer_left = footer_left
    _C._watermark = watermark

    doc.build(story, canvasmaker=_C)
    return path_or_buffer


# ============================================================================
# SECTION 5 — FILENAMES + BULK (per-batch) GENERATION (spec Section 11)
# ============================================================================

def safe_filename(result, ext):
    """A tidy, unique-enough filename from the offering identity, e.g.
    'E02_Y3_CSE23CT301_Prof-Manoj.xlsx'. Non-filename characters are stripped."""
    o = result["offering"]
    # Include the offering id (when present) so two offerings that share a
    # dept/year/code/faculty — or a batch that repeats one — never collide on
    # the same archive filename.
    oid = o["id"] if ("id" in o.keys()) else None
    # For a pooled elective the anchor's single dept_code would be misleading in the
    # filename, so we label it 'ELECTIVE' (the report itself lists every programme).
    dept_part = "ELECTIVE" if result.get("is_consolidated") else (o["dept_code"] or "NA")
    parts = [dept_part, f"Y{o['year_of_study']}",
             o["course_code"] or "NOCODE", (o["faculty"] or "NoFaculty")]
    if oid is not None:
        parts.append(f"id{oid}")
    raw = "_".join(str(p) for p in parts)
    keep = "".join(ch if (ch.isalnum() or ch in "._-") else "-" for ch in raw)
    return f"{keep}.{ext}"


def build_batch_excel_zip(results, zip_path):
    """Bulk Excel: write one .xlsx per offering into a single .zip (spec 11).

    `results` is a list of scored-offering dicts. We render each to an in-memory
    workbook and add it to the archive, so a whole batch downloads as one file.
    """
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for res in results:
            buf = io.BytesIO()
            build_excel_report(res, buf)
            zf.writestr(safe_filename(res, "xlsx"), buf.getvalue())
    return zip_path


def build_batch_pdf(results, pdf_path):
    """Bulk PDF: one combined, multi-page PDF for the whole batch (spec 11).

    Each offering's report is built into its own PDF page-set in memory, then
    the pages are concatenated with reportlab's low-level canvas import via
    pypdf-free merging: we render every offering into ONE SimpleDocTemplate by
    concatenating their stories with a PageBreak between them.
    """
    from reportlab.platypus import PageBreak
    ss = _pdf_styles()
    doc = SimpleDocTemplate(pdf_path, pagesize=A4,
                            topMargin=14 * mm, bottomMargin=14 * mm,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            title="SRET Feedback Reports (Batch)")
    story = []
    for idx, res in enumerate(results):
        # Reuse the single-report builder's story by capturing it: simplest is to
        # build each to a temp buffer then merge PDFs, but to avoid a PDF-merge
        # dependency we instead re-run the layout inline here via a shared story
        # accumulator. We call a private helper that APPENDS to `story`.
        _append_pdf_story(res, story, ss)
        # Each report still starts on a fresh page (PageBreak between reports), but
        # (v3.2) no blank padding is added — reports print at their true length.
        if idx != len(results) - 1:
            story.append(PageBreak())
    # Use the numbered canvas so the combined PDF carries the confidential footer
    # and "Page X of Y" across all reports.
    watermark = results[0].get("watermark") if results and hasattr(results[0], "get") else None

    class _C(NumberedCanvas):
        pass
    _C.footer_left = "SRET Feedback (batch)"
    _C._watermark = watermark
    doc.build(story, canvasmaker=_C)
    return pdf_path


def _append_pdf_story(result, story, ss):
    """Append one offering's full report flowables to a shared `story` list.

    This is the same layout as build_pdf_report, factored so the combined-batch
    PDF can stack many offerings into one document without merging PDF files.
    """
    from reportlab.platypus import PageBreak, Image as RLImage  # local imports
    vm = build_view_model(result)
    # TITLE: the college banner image (as seen on all our documents), then just
    # "Course Feedback Report" — no version number, no "SRET" text. The banner
    # lives at app/static/banner.png; if it is somehow missing we fall back to a
    # plain text title so a report is never blank.
    banner_path = os.path.join(Config.BASE_DIR, "static", "banner.png")
    if os.path.exists(banner_path):
        # Banner art is 2000x400 (5:1). Fit it to the content width (~180mm) and
        # keep the aspect ratio so it never distorts.
        banner = RLImage(banner_path, width=180 * mm, height=36 * mm)
        banner.hAlign = "CENTER"
        story.append(banner)
        story.append(Spacer(1, 6))
    story.append(Paragraph("<b>Course Feedback Report</b>", ss["Title"]))
    story.append(Spacer(1, 6))
    h = vm["header"]
    ident_lines = (
        f"<b>ACADEMIC YEAR:</b> {h['academic_year']} &nbsp;&nbsp; "
        f"<b>YEAR / TERM:</b> {h['year_term']}"
        + (f" &nbsp;&nbsp; <b>SECTION:</b> {h['section']}" if h['section'] else "")
        + "<br/>"
        + f"<b>COURSE CODE:</b> {h['course_code']} &nbsp;&nbsp; "
        f"<b>COURSE NAME:</b> {h['course_name']}<br/>"
        f"<b>PROGRAM CODE:</b> {h['dept_code']} &nbsp;&nbsp; "
        f"<b>PROGRAMME:</b> {h['programme']}<br/>"
        f"<b>FACULTY NAME:</b> {h['faculty']}<br/>"
        f"<b>CATEGORY:</b> {h['category']} — {h['category_name']}"
    )
    story.append(Paragraph(ident_lines, ss["Ident"]))
    story.append(Spacer(1, 8))
    box = Table([
        ["No. of Students Feedback Recorded", str(vm["n_recorded"])],
        ["OVERALL SCORE (out of 10)", _fmt(vm["overall"])],
    ], colWidths=[110 * mm, 60 * mm])
    box.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EEF1F4")),
        ("BACKGROUND", (1, 1), (1, 1), colors.HexColor("#0B3D68")),
        ("TEXTCOLOR", (1, 1), (1, 1), colors.white),
        ("FONTSIZE", (1, 1), (1, 1), 16),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C9CED3")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # The overall score is the single most important number in the report, so
        # centre it in its blue cell BOTH ways. Without an explicit ALIGN it was
        # left-aligned (hugging the grid line).
        ("ALIGN", (1, 1), (1, 1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        # Vertical fix: VALIGN MIDDLE reserves the 16pt font's descender space, so
        # the big digits sit ~4pt low relative to the label. Asymmetric padding
        # (less on top, more on bottom) raises them to the same optical centre as
        # the label text — measured level to within 0.2pt.
        ("TOPPADDING", (1, 1), (1, 1), 3),
        ("BOTTOMPADDING", (1, 1), (1, 1), 11),
    ]))
    story.append(box)
    story.append(Spacer(1, 10))
    sec_data = [["Section", "Score /10"]]
    for srow in vm["section_rows"]:
        sec_data.append([srow["title"], _fmt(srow["score"])])
    sec_tbl = Table(sec_data, colWidths=[130 * mm, 40 * mm])
    sec_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3D68")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C9CED3")),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(sec_tbl)
    story.append(Spacer(1, 12))
    for sec in vm["detail_sections"]:
        if not sec["questions"]:
            continue
        block = [Paragraph(sec["title"], ss["SecTitle"])]
        labels = [lbl for (lbl, _c) in sec["questions"][0]["counts"]]
        # A leading "#" column numbers the questions Q1..Qn; the trailing "Avg"
        # column carries each question's average /10 — the exact numbers that used
        # to be drawn as a bar chart, kept as data after the v3.2 chart removal.
        # HEADER CELLS ARE PARAGRAPHS, not raw strings: a plain string does NOT
        # wrap inside a narrow ReportLab table cell, so long option labels like
        # "Strongly Disagree" used to spill over and overlap the neighbouring
        # columns ("Strongly DisagreAevg"). Wrapping each label in a Paragraph
        # with the centred header style lets it break onto two lines and stay
        # inside its own column.
        header_cells = ([Paragraph("#", ss["Hdr"]), Paragraph("Question", ss["Hdr"])]
                        + [Paragraph(lbl, ss["Hdr"]) for lbl in labels]
                        + [Paragraph("Avg", ss["Hdr"])])
        tdata = [header_cells]
        # A section that scores at section level (Syllabus) leaves per-question
        # averages None; for such a SINGLE-question section, show the section score
        # in the Avg cell instead of a blank. Multi-question sections keep None.
        sec_score = sec.get("score")
        single_q = len(sec["questions"]) == 1
        for idx, qb in enumerate(sec["questions"], start=1):
            qn = f"Q{idx}"
            disp_avg = qb["average"]
            if disp_avg is None and single_q:
                disp_avg = sec_score
            count_map = dict(qb["counts"])
            counts = [count_map.get(lbl, 0) for lbl in labels]
            tdata.append([qn, Paragraph(qb["text"], ss["Ident"])]
                         + [str(c) for c in counts] + [_fmt(disp_avg)])
        n_opt = len(labels)
        num_w = 8 * mm
        # Widths total unchanged (# 8 + question 70 + options 100 + avg 14 = 192).
        tbl = Table(
            tdata,
            colWidths=[num_w, 70 * mm] + [(100 * mm) / max(n_opt, 1)] * n_opt + [14 * mm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF1F4")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),   # bold Q# so it stands out
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#C9CED3")),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("ALIGN", (0, 1), (0, -1), "CENTER"),              # centre the Q# column
            ("ALIGN", (2, 1), (-1, -1), "CENTER"),             # centre the count/avg columns
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        block.append(tbl)
        block.append(Spacer(1, 10))
        story.append(KeepTogether(block))
    # (v3.2) The open-ended comments block and the department-code legend that
    # used to close each report were removed as part of the report slim-down.
