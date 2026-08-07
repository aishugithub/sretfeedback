# ============================================================================
# allocation_rigid.py  —  Rigid course-allocation upload (spec v3 §7.2)
# ============================================================================
# WHERE THIS FITS IN THE WHOLE APPLICATION
# ----------------------------------------------------------------------------
# This is the SECOND of the two uploads that begin every cycle — the COURSE
# ALLOCATION ("what"): which faculty teaches which course to which group. In v1
# an auto-converter tried to unpivot the college's merged-cell matrix workbook
# and GUESSED at elective baskets; v3 replaces that with a RIGID, flat template
# the admin fills — one row per teaching assignment, no merged cells, no
# multi-value cells. Converting the college's working sheet into this template is
# explicitly the admin's job (§7.2), because a merged "Elective-I: A/B/C" beside
# three faculty names is irrecoverably ambiguous to any parser.
#
# Each committed row becomes an `offering` (the ATOMIC teaching assignment, §6)
# with a system-generated id the admin never types. The importer validates
# EVERYTHING before writing (a dry run), including the crucial cross-check that
# every (programme, year, section) it names actually exists in the cycle roster
# (§7.2) — this is what catches an omitted roster row loudly.
#
# (v3.1) The old matrix converter (allocation_importer.py) and its "Re-import"
# button have been RETIRED — this rigid, explicit-category importer is now the
# only allocation path.
# ----------------------------------------------------------------------------

import re
from openpyxl import load_workbook, Workbook

from config import Config


# The rigid template columns, in order (spec §7.2).
#
# v2.0 · Module 5 CHANGE — FACULTY DECOUPLING. The `faculty_email` column is
# GONE. Email, phone and home department are now owned by the Faculty Master
# (the `faculty` table / the "Manage Faculty" admin page) and looked up by join
# on `faculty_id` (the employee/staff number). The allocation file therefore
# carries only the two faculty facts a human needs to read/verify a row:
#   * faculty_id   — the employee number, the KEY that links to the Faculty Master
#   * faculty_name — for human readability only (so you can eyeball the sheet)
# Everything else about the person lives in one place and never has to be re-typed
# into the allocation every cycle.
#
# ROBUSTNESS: parse_allocation now maps columns BY HEADER NAME, not by fixed
# position (see below). So adding/removing/reordering a column — like dropping
# faculty_email — can never silently misalign the data, and an older file that
# still contains a faculty_email column simply has it ignored.
ALLOC_HEADERS = ["academic_year", "semester", "programme_code", "year_of_study",
                 "section", "course_code", "course_title", "course_type",
                 "feedback_category",  # THEORY/LAB/SKILL/AE — selects the questionnaire
                 "elective_basket", "faculty_name", "faculty_id",
                 "role", "expected_students"]

# The four legal course types (spec §7.2, §7.4). This is an ENROLLMENT flag only
# (electives derive their headcount from the enrollment upload; everything else
# uses expected_students). It does NOT choose the feedback questionnaire.
COURSE_TYPES = {"CORE", "ELECTIVE", "LAB", "PROJECT"}

# The legal feedback categories are NOT hardcoded here. They are whatever rows
# currently exist in master.db's `category` table — seeded as THEORY/LAB/SKILL/AE
# (seed_data.py), but EXTENSIBLE at runtime: an admin can add e.g. "INTERNSHIP"
# via Admin -> Categories (config_routes.categories_new) WITHOUT any code change,
# and the upload validator below will accept that new code automatically because
# it reads the set fresh from the DB on every validation (see valid_categories).
# The admin DECLARES this per row instead of the system reverse-engineering it
# from the course code, because a code/name is an unreliable source of truth for
# pedagogy (spec §3).

# A permissive but real email pattern for faculty-email validation.
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _norm(s):
    """Trim a cell to a clean string; None/blank -> ''."""
    return ("" if s is None else str(s)).strip()


# NOTE (v3.1): the old _detect_category_id() helper was REMOVED. It guessed the
# feedback category from the two-letter segment of the course code
# (CSE23[CT]201 -> 'CT' -> THEORY). That derivation was fragile — any code that
# didn't fit the ABC99XX9 shape, or used a segment outside the four-entry map,
# silently produced category_id = NULL and an "uncategorised" offering. The
# category is now a REQUIRED, explicit column in the upload (feedback_category),
# so there is nothing left to auto-detect. Config.CATEGORY_BY_CODE_SEGMENT is no
# longer consulted on this path.


# ----------------------------------------------------------------------------
# parse_allocation(path) — read the rigid template's first (or 'Allocation')
# sheet into raw row dicts, each tagged with its 1-based Excel row number.
# ----------------------------------------------------------------------------
def parse_allocation(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = None
    for w in wb.worksheets:
        if (w.title or "").strip().lower().startswith(("allocation", "sheet")):
            ws = w
            break
    if ws is None:
        ws = wb.worksheets[0]

    rows = []
    header_map = None       # {normalised column name -> its 0-based index}
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row is None or all(_norm(c) == "" for c in row):
            continue

        # The FIRST non-blank row is the header. We build a name->index map from
        # it so every field below is read BY NAME. This makes the parser immune
        # to column reordering and to the removal of faculty_email (Module 5):
        # a column that isn't present just resolves to "".
        if header_map is None:
            header_map = {}
            for ci, val in enumerate(row):
                key = _norm(val).lower()
                if key:
                    header_map[key] = ci
            continue

        # cell(name) — fetch this row's value for a named column, or "" if the
        # column is absent or the row is short.
        def cell(name, _row=row):
            ci = header_map.get(name)
            if ci is None or ci >= len(_row):
                return ""
            return _norm(_row[ci])

        rows.append({
            "excel_row": idx,
            "academic_year": cell("academic_year"),
            "semester": cell("semester").upper(),
            "programme_code": cell("programme_code").upper(),
            "year_raw": cell("year_of_study"),
            "section": (cell("section") or "NA").upper(),
            "course_code": cell("course_code"),
            "course_title": cell("course_title"),
            "course_type": cell("course_type").upper() or "CORE",
            # REQUIRED: a blank stays "" and is REJECTED by validate_allocation.
            "feedback_category": cell("feedback_category").upper(),
            "elective_basket": cell("elective_basket") or None,
            # Faculty: only the emp-no (the key) + the name (readability). Email is
            # resolved from the Faculty Master in validate_allocation, not read here.
            "faculty_name": cell("faculty_name"),
            "faculty_id": cell("faculty_id"),
            "role": cell("role") or "Primary",
            "expected_raw": cell("expected_students"),
        })
    wb.close()
    return rows


# ----------------------------------------------------------------------------
# validate_allocation(rows, master, cycle_code) — the dry run. Runs every §7.2
# validation and returns (clean, errors, warnings, recon). Nothing is written on
# any hard error.
#
# Validations (spec §7.2):
#   * programme_code exists in the programme master                    -> error
#   * (programme, year, section) exists in THIS cycle's roster         -> error
#     (the cross-check that catches an omitted roster row)
#   * course_type is CORE/ELECTIVE/LAB/PROJECT                         -> error
#   * elective_basket present iff course_type = ELECTIVE               -> error
#   * no duplicate (course_code, faculty_id, section) row              -> error
#   * faculty_id exists in the Faculty Master (email resolved from it)  -> warning
#   * the same faculty_id never maps to two different names            -> warning
# ----------------------------------------------------------------------------
def validate_allocation(rows, master, cycle_code):
    errors, warnings, clean = [], [], []

    valid_progs = {r["code"] for r in master.execute("SELECT code FROM programme")}

    # The legal feedback categories = whatever is configured in the DB right now.
    # Read fresh each run so admin-added categories (e.g. INTERNSHIP) are honoured
    # with zero code changes. Sorted list is only for a friendly error message.
    valid_categories = {r["code"] for r in master.execute("SELECT code FROM category")}
    category_choices = "/".join(sorted(valid_categories)) or "(none configured)"

    # The set of (programme, year, section) audiences present in the cycle roster,
    # used for the cross-check. Built from roster_range so it works even before
    # the roster is expanded into students.
    roster_groups = set()
    for rr in master.execute(
            "SELECT programme_code, year_of_study, section FROM roster_range "
            "WHERE cycle_code = ?", (cycle_code,)):
        roster_groups.add((rr["programme_code"], rr["year_of_study"], rr["section"]))

    seen_assign = set()      # (course_code, faculty_id, section) dedupe
    faculty_names = {}       # faculty_id -> first name seen (typo catcher)

    # Faculty Master (v2.0 · Module 5): the set of known employee numbers and an
    # emp-no -> email lookup. The allocation now carries only the emp-no, so the
    # email that the result/ATR mail needs is resolved HERE from the master (the
    # single source of truth). An emp-no the master doesn't know is flagged (soft)
    # so the admin adds the person on the Manage Faculty page. Guarded so a DB that
    # predates the faculty table (before the Module 5 migration) still validates.
    faculty_email_by_emp, known_emp = {}, set()
    try:
        for fr in master.execute("SELECT emp_no, email FROM faculty"):
            known_emp.add(fr["emp_no"])
            if fr["email"]:
                faculty_email_by_emp[fr["emp_no"]] = fr["email"]
    except Exception:
        pass   # no faculty table yet -> treat as "master empty" (all rows warn)

    for r in rows:
        rn = r["excel_row"]
        prog = r["programme_code"]

        if prog not in valid_progs:
            errors.append(f"Row {rn}: programme_code '{prog}' is not in the "
                          f"programme master.")

        year = None
        if not r["year_raw"].isdigit() or not (1 <= int(r["year_raw"]) <= 4):
            errors.append(f"Row {rn}: year_of_study '{r['year_raw']}' must be 1-4.")
        else:
            year = int(r["year_raw"])

        ctype = r["course_type"]
        if ctype not in COURSE_TYPES:
            errors.append(f"Row {rn}: course_type '{ctype}' must be one of "
                          f"CORE/ELECTIVE/LAB/PROJECT.")

        # feedback_category is STRICTLY REQUIRED and picks the questionnaire.
        # A blank ("") or any value outside the four legal codes is a HARD ERROR,
        # so no offering can ever reach the DB without a valid template selector.
        fcat = r["feedback_category"]
        if fcat not in valid_categories:
            errors.append(f"Row {rn}: feedback_category '{fcat}' must be one of "
                          f"{category_choices} (this selects the feedback form). "
                          f"Add new categories under Admin -> Categories first.")

        # elective_basket present iff ELECTIVE
        if ctype == "ELECTIVE" and not r["elective_basket"]:
            errors.append(f"Row {rn}: ELECTIVE course requires an elective_basket.")
        if ctype != "ELECTIVE" and r["elective_basket"]:
            warnings.append(f"Row {rn}: elective_basket set on a non-elective "
                            f"course — ignored.")

        if not r["course_code"]:
            errors.append(f"Row {rn}: course_code is blank.")
        if not r["faculty_name"]:
            errors.append(f"Row {rn}: faculty_name is blank.")
        if not r["faculty_id"]:
            errors.append(f"Row {rn}: faculty_id is blank.")

        # Cross-check against the roster (only meaningful for non-electives, whose
        # audience is the class; electives resolve by enrollment so a roster group
        # match is not required, though the enrolled students must be in the roster).
        if year is not None and prog in valid_progs and ctype != "ELECTIVE":
            if (prog, year, r["section"]) not in roster_groups:
                errors.append(f"Row {rn}: ({prog}, year {year}, section "
                              f"{r['section']}) is not in the cycle roster — "
                              f"upload/adjust the roster first.")

        # Duplicate assignment guard.
        # An offering is uniquely identified by WHO is taught (programme + year +
        # section) as well as WHAT/ by WHOM (course_code + faculty_id). The same
        # course code, taught by the same faculty, to the same section LABEL but in
        # a DIFFERENT programme or a DIFFERENT year is a legitimately separate
        # offering — so programme_code and year_of_study MUST be part of the key.
        # (The old key was just (course_code, faculty_id, section), which wrongly
        # merged those distinct assignments and reported false duplicates.)
        key = (prog, year, r["section"], r["course_code"], r["faculty_id"])
        if key in seen_assign:
            errors.append(
                f"Row {rn}: duplicate assignment "
                f"(programme {prog}, year {year}, section {r['section']}, "
                f"course {r['course_code']}, faculty {r['faculty_id']}).")
        seen_assign.add(key)

        # Faculty Master membership (soft). The email lives in the master now, so
        # instead of validating an email column (there isn't one) we check that the
        # emp-no is known. If it isn't, results/ATR mail can't be addressed until
        # the person is added on Manage Faculty — a warning, not a hard error, so
        # the allocation still imports (e.g. a brand-new hire added to the master
        # a moment later).
        if r["faculty_id"] and r["faculty_id"] not in known_emp:
            warnings.append(f"Row {rn}: faculty_id '{r['faculty_id']}' "
                            f"({r['faculty_name']}) is not in the Faculty Master — "
                            f"add them under Manage Faculty so their email and "
                            f"department resolve.")

        # Same faculty_id, two names -> likely a typo (soft)
        if r["faculty_id"]:
            if r["faculty_id"] in faculty_names and faculty_names[r["faculty_id"]] != r["faculty_name"]:
                warnings.append(f"Row {rn}: faculty_id '{r['faculty_id']}' maps to "
                                f"both '{faculty_names[r['faculty_id']]}' and "
                                f"'{r['faculty_name']}'.")
            else:
                faculty_names.setdefault(r["faculty_id"], r["faculty_name"])

        # A row is "clean" only if every hard requirement holds — now including a
        # valid feedback_category, so the questionnaire selector is guaranteed
        # present on every row we carry forward to commit.
        if year is not None and ctype in COURSE_TYPES \
                and fcat in valid_categories and prog in valid_progs \
                and r["course_code"] and r["faculty_id"]:
            clean.append({
                "cycle_code": cycle_code,
                "academic_year": r["academic_year"] or None,
                "semester": r["semester"] or None,
                "programme": Config.PROGRAMMES.get(prog, (None, None))[1],
                "dept_code": prog,
                "year_of_study": year,
                "section": r["section"],
                "course_code": r["course_code"],
                "course_name": r["course_title"],
                "course_type": ctype,
                # The declared questionnaire code, carried to commit_allocation,
                # which resolves it to a category_id.
                "feedback_category": fcat,
                "elective_basket": r["elective_basket"] if ctype == "ELECTIVE" else None,
                "faculty": r["faculty_name"],
                # Email is taken from the Faculty Master by emp-no (single source
                # of truth), NOT from the allocation file. None when the faculty is
                # not in the master yet — the send path simply skips them (and the
                # warning above told the admin to add them).
                "faculty_email": faculty_email_by_emp.get(r["faculty_id"]),
                "faculty_id": r["faculty_id"],
                "role": r["role"],
                "expected_students": int(r["expected_raw"]) if r["expected_raw"].isdigit() else None,
            })

    if errors:
        clean = []

    recon = _reconcile(clean)
    return clean, errors, warnings, recon


def _reconcile(clean):
    """Summary the admin confirms before commit (spec §7.2 reconciliation)."""
    courses, faculty, electives = set(), set(), set()
    for o in clean:
        courses.add(o["course_code"])
        faculty.add(o["faculty_id"])
        if o["course_type"] == "ELECTIVE":
            electives.add((o["course_code"], o["faculty_id"]))
    return {
        "assignments": len(clean),
        "distinct_courses": len(courses),
        "distinct_faculty": len(faculty),
        "elective_assignments": len(electives),
    }


# ----------------------------------------------------------------------------
# commit_allocation(clean, master, cycle_code) — replace this cycle's allocation
# with the validated rows (a re-upload is authoritative). The feedback category
# is taken DIRECTLY from the row's declared feedback_category (no code guessing);
# validate_allocation already guaranteed it is one of THEORY/LAB/SKILL/AE.
# Idempotent.
# ----------------------------------------------------------------------------
def commit_allocation(clean, master, cycle_code):
    # {category code -> id} lookup from the seeded category table. Since the four
    # legal feedback_category codes are exactly these rows' codes, every clean row
    # resolves to a real id here.
    cat_rows = master.execute("SELECT id, code FROM category").fetchall()
    category_id_by_code = {r["code"]: r["id"] for r in cat_rows}

    cur = master.cursor()
    # Removing prior offerings for the cycle also clears their elective
    # enrollments, which would otherwise dangle.
    old_ids = [r["id"] for r in cur.execute(
        "SELECT id FROM offering WHERE cycle_code = ?", (cycle_code,))]
    if old_ids:
        cur.executemany("DELETE FROM enrollment WHERE offering_id = ?",
                        [(i,) for i in old_ids])
    cur.execute("DELETE FROM offering WHERE cycle_code = ?", (cycle_code,))

    for o in clean:
        # Resolve the ADMIN-DECLARED category code to its id. No fallback: the
        # value was validated, so a missing key would be a genuine data-integrity
        # bug rather than a routine "couldn't guess" case.
        cat_id = category_id_by_code[o["feedback_category"]]
        cur.execute(
            "INSERT INTO offering (cycle_code, academic_year, semester, programme, "
            "dept_code, year_of_study, section, course_code, course_name, "
            "course_type, elective_basket, faculty, faculty_email, faculty_id, "
            "role, expected_students, category_id) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (o["cycle_code"], o["academic_year"], o["semester"], o["programme"],
             o["dept_code"], o["year_of_study"], o["section"], o["course_code"],
             o["course_name"], o["course_type"], o["elective_basket"], o["faculty"],
             o["faculty_email"], o["faculty_id"], o["role"], o["expected_students"],
             cat_id),
        )
    master.commit()
    return len(clean)


def import_allocation_rigid(path, master, cycle_code, commit=False):
    """The one call the admin route makes. commit=False is the dry run."""
    rows = parse_allocation(path)
    clean, errors, warnings, recon = validate_allocation(rows, master, cycle_code)
    inserted = 0
    if commit and not errors:
        inserted = commit_allocation(clean, master, cycle_code)
    return {
        "rows": len(rows),
        "assignments": len(clean),
        "inserted": inserted,
        "committed": bool(commit and not errors),
        "errors": errors,
        "warnings": warnings,
        "reconciliation": recon,
    }


# ----------------------------------------------------------------------------
# build_allocation_template(path, cycle_label) — the downloadable rigid template
# (spec §7.2). Sheet 1 = headers + worked example rows (incl. an elective basket
# taught by two faculty as two rows); Sheet 2 = instructions.
# ----------------------------------------------------------------------------
def build_allocation_template(path, cycle_label="AY 2026-27 CA1"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Allocation"
    ws.append(ALLOC_HEADERS)
    # Worked examples. The 9th value in every row is feedback_category — the
    # questionnaire selector. NOTE (Module 5): there is NO faculty_email column —
    # the faculty is identified by faculty_id (their employee number) and shown by
    # faculty_name for readability; email/phone/department live in Manage Faculty.
    # Core theory course -> THEORY questionnaire.
    ws.append(["2026-27", "ODD", "E02", 2, "A", "CSE23CT201",
               "Computer Organization and Architecture", "CORE", "THEORY", "",
               "Dr. Priya R", "16093", "Primary", 30])
    # Lab course -> LAB questionnaire.
    ws.append(["2026-27", "ODD", "E02", 2, "A", "CSE23CL202",
               "Operating Systems Lab", "LAB", "LAB", "",
               "Dr. Kumar S", "16094", "Primary", 30])
    # Two-faculty elective basket (two rows) — elective, but a theory questionnaire.
    ws.append(["2026-27", "ODD", "E02", 4, "A", "CS4001", "Cloud Computing",
               "ELECTIVE", "THEORY", "Elective-I", "Dr. Anand", "16445", "Primary", ""])
    ws.append(["2026-27", "ODD", "E02", 4, "A", "CS4001", "Cloud Computing",
               "ELECTIVE", "THEORY", "Elective-I", "Dr. Latha", "16191", "Primary", ""])

    ins = wb.create_sheet("Instructions")
    for line in [
        [f"Course allocation template — {cycle_label}"],
        [""],
        ["ONE ROW PER TEACHING ASSIGNMENT. Flat. No merged cells."],
        ["A course taught by 3 faculty = 3 rows. An elective basket with 5 options"],
        ["= 5 rows, each with its own faculty."],
        [""],
        ["course_type     : CORE / ELECTIVE / LAB / PROJECT. (Enrollment flag only —"],
        ["                  decides how the class size is counted; NOT the feedback form.)"],
        ["feedback_category: REQUIRED. One of the categories configured in the app"],
        ["                  (default THEORY / LAB / SKILL / AE; add more under"],
        ["                  Admin -> Categories, e.g. INTERNSHIP). This decides which"],
        ["                  feedback questionnaire the students receive. There is no"],
        ["                  auto-detection — every row must state it, or it is rejected."],
        ["elective_basket : required ONLY when course_type = ELECTIVE (e.g. Elective-I)."],
        ["section        : A / B ... or NA if the class is not divided."],
        ["--- FACULTY (changed in v2.0) --------------------------------------------"],
        ["faculty_id     : the EMPLOYEE / STAFF NUMBER (e.g. 16093). This is the key"],
        ["                  that links the course to the person. It MUST match a"],
        ["                  record in Admin -> Manage Faculty."],
        ["faculty_name   : for readability only, so you can eyeball the sheet. The"],
        ["                  app takes the real email, phone and DEPARTMENT from the"],
        ["                  Faculty Master, NOT from this file."],
        ["                  (There is deliberately NO faculty_email column any more —"],
        ["                  change an email once, in Manage Faculty, not every cycle.)"],
        ["                  If an emp-no here is not yet in Manage Faculty, the row"],
        ["                  still imports but is flagged so you can add the person."],
        ["expected_students: optional; leave BLANK for electives (derived from enrollment)."],
        [""],
        ["Every (programme_code, year_of_study, section) for a CORE/LAB/PROJECT row"],
        ["must already exist in the uploaded roster — otherwise the row is rejected."],
        ["Electives resolve by the separate enrollment upload, not by the roster."],
    ]:
        ins.append(line)

    wb.save(path)
    return path
